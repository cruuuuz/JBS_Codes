import openpyxl
import pandas


class Cwt_per_usd:
    @staticmethod
    def convert_cwt_per_usd(path: str):
        wb_base = openpyxl.load_workbook(f'{path}/Base_Cattle_Prices.xlsx')
        ws_base = wb_base.active

        wb_currency = openpyxl.load_workbook(f'{path}/Base_Currency.xlsx')
        ws_currency = wb_currency['Hist']

        wb_yield_uruguay = openpyxl.load_workbook(f'{path}/Preços_Uruguai.xlsx')
        ws_yield_uruguay = wb_yield_uruguay['Sheet1']

        line = 1
        for i in range (1, 10000):
            try:
                ws_base[f"B{line+1}"] = ws_base[f"B{line+1}"].value #USA
                ws_base[f"C{line+1}"] = ws_base[f"C{line+1}"].value #AUS
                ws_base[f"D{line+1}"] = ws_base[f"D{line+1}"].value #BRA
                ws_base[f"E{line+1}"] = ws_base[f"E{line+1}"].value*(ws_yield_uruguay[f"C{line+1}"].value/100) / 2.2046 * 100   #Converte Uruguay
                ws_base[f"G{line+1}"] = ws_base[f"G{line+1}"].value*ws_currency[f"I{line+1}"].value * 1.025                     #Converte Canada
                ws_base[f"H{line+1}"] = ws_base[f"H{line+1}"].value*ws_currency[f"E{line+1}"].value / 2.2046 * 100              #Converte Argentina
                ws_base[f"I{line+1}"] = ws_base[f"I{line+1}"].value*ws_currency[f"L{line+1}"].value / 100 / 2.2046 * 100 * 0.60 #Converte Reino Unido
            except:
                ws_base[f"B{line+1}"] = ws_base[f"B{line+1}"].value
                ws_base[f"C{line+1}"] = ws_base[f"C{line+1}"].value
                ws_base[f"D{line+1}"] = ws_base[f"D{line+1}"].value
                ws_base[f"E{line+1}"] = ws_base[f"E{line+1}"].value
                ws_base[f"G{line+1}"] = ws_base[f"G{line+1}"].value
                ws_base[f"H{line+1}"] = f"=H{line+2}"
                ws_base[f"I{line+1}"] = ws_base[f"I{line+1}"].value


                # ws_base[f"F{line+1}"] = f"=F{line+2}"
            line += 1

        df = pandas.DataFrame(ws_base.values)
        df = df.drop(columns=[9])
        df = df.drop(columns=[10])
        df = df.drop(columns=[11])


        df.columns = df.iloc[0,:]
        df = df.drop(df.index[0])
        df[f'{df.columns[1]}'] = pandas.to_numeric(df[f'{df.columns[1]}'], errors='coerce')
        df[f'{df.columns[2]}'] = pandas.to_numeric(df[f'{df.columns[2]}'], errors='coerce')
        df[f'{df.columns[3]}'] = pandas.to_numeric(df[f'{df.columns[3]}'], errors='coerce')
        df[f'{df.columns[4]}'] = pandas.to_numeric(df[f'{df.columns[4]}'], errors='coerce')
        df[f'{df.columns[5]}'] = pandas.to_numeric(df[f'{df.columns[5]}'], errors='coerce')
        df[f'{df.columns[6]}'] = pandas.to_numeric(df[f'{df.columns[6]}'], errors='coerce')
        df[f'{df.columns[7]}'] = pandas.to_numeric(df[f'{df.columns[7]}'], errors='coerce')
        df[f'{df.columns[8]}'] = pandas.to_numeric(df[f'{df.columns[8]}'], errors='coerce')
        df[f'{df.columns[1]}'] = df[f'{df.columns[1]}'].replace(0,df.iloc[1,1])
        df[f'{df.columns[2]}'] = df[f'{df.columns[2]}'].replace(0,df.iloc[1,2])
        df[f'{df.columns[3]}'] = df[f'{df.columns[3]}'].replace(0,df.iloc[1,3])
        df[f'{df.columns[4]}'] = df[f'{df.columns[4]}'].replace(0,df.iloc[1,4])
        df[f'{df.columns[5]}'] = df[f'{df.columns[5]}'].replace(0,df.iloc[1,5])
        df[f'{df.columns[6]}'] = df[f'{df.columns[6]}'].replace(0,df.iloc[1,6])
        df[f'{df.columns[7]}'] = df[f'{df.columns[7]}'].replace(0,df.iloc[1,7])
        df[f'{df.columns[8]}'] = df[f'{df.columns[8]}'].replace(0,df.iloc[1,8])

        df.loc[-1] = df.columns.tolist()
        df.index = df.index + 1
        df = df.sort_index()
        df.columns = [0,1,2,3,4,5,6,7,8]
        df = df.fillna(method='bfill')

        df_cwt = df.head(2000)
        df_cwt = df_cwt.drop(df.index[0])
        df_cwt.columns = ["Dates", "USA - CWT/U$", "AUSTRALIA - AUD/KG", "BRAZIL - R$/@", "URUGUAY - U$/KG", "FRANCE - EUR/KG", "CANADA - CWT/CAD", "ARGENTINA - CAB/ARS", "GREAT_BRITAIN - POUNDS/KG"]
        
        df_cwt.to_excel(f'{path}/Global_Cattle_Prices.xlsx', index=False)
        print('Convertion to CWT/USD is done')