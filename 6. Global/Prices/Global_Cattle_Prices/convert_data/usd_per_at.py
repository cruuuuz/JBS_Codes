import openpyxl
import pandas

class Usd_per_at:
    @staticmethod
    def convert_usd_per_at(path: str):
        wb_base = openpyxl.load_workbook(f'{path}/Base_Cattle_Prices.xlsx')
        ws_base = wb_base.active

        wb_currency = openpyxl.load_workbook(f'{path}/Base_Currency.xlsx')
        ws_currency = wb_currency['Hist']

        wb_yield_uruguay = openpyxl.load_workbook(f'{path}/Preços_Uruguai.xlsx')
        ws_yield_uruguay = wb_yield_uruguay['Sheet1']

        line = 1
        for i in range (1, 10000):
            try:
                ws_base[f"B{line+1}"] = ws_base[f"B{line+1}"].value/100 * 2.2046 * 15 / 0.63 #Converte USA
                ws_base[f"C{line+1}"] = ws_base[f"C{line+1}"].value/100 * 2.2046 * 15 / 0.58 #Converte AUS
                ws_base[f"D{line+1}"] = ws_base[f"D{line+1}"].value/100 * 2.2046 * 15 / 0.55 #Converte Brazil
                ws_base[f"E{line+1}"] = ws_base[f"E{line+1}"].value * 15                     #Converte Uruguay
                ws_base[f"G{line+1}"] = ws_base[f"G{line+1}"].value*ws_currency[f"I{line+1}"].value * 1.025 /100 * 2.2046 * 15 / 0.63 #Converte CAN
                ws_base[f"H{line+1}"] = ws_base[f"H{line+1}"].value*ws_currency[f"E{line+1}"].value * 15 / 0.55                       #Converte Argentina
                ws_base[f"I{line+1}"] = ws_base[f"I{line+1}"].value*ws_currency[f"L{line+1}"].value / 100 * 15                        #Converte Reino Unido
            except:
                ws_base[f"B{line+1}"] = ws_base[f"B{line+1}"].value
                ws_base[f"C{line+1}"] = ws_base[f"C{line+1}"].value
                ws_base[f"D{line+1}"] = ws_base[f"D{line+1}"].value
                ws_base[f"E{line+1}"] = ws_base[f"E{line+1}"].value
                ws_base[f"G{line+1}"] = ws_base[f"G{line+1}"].value
                ws_base[f"H{line+1}"] = f"=H{line+2}"
                ws_base[f"I{line+1}"] = ws_base[f"I{line+1}"].value
            line += 1

        df = pandas.DataFrame(ws_base.values)
        df = df.drop(columns=[9])
        df = df.drop(columns=[10])
        df = df.drop(columns=[11])
        df = df.drop(0)

        df.columns = ['Dates', 'USA - CWT/U$', 'AUSTRALIA - AUD/KG', 'BRAZIL - R$/@', 'URUGUAY - U$/KG', 'FRANCE - EUR/KG', 'CANADA - CWT/CAD', 'ARGENTINA - CAB/ARS', "GREAT_BRITAIN - POUNDS/KG"]


        df.to_excel(f'{path}/Global_Cattle_USD_Arroba.xlsx', index=False)
        print('Convertion to USD/@ is done')
