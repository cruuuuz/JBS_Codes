import requests
import pandas
import openpyxl
import sys
import datetime
from pathlib import Path
from datetime import date
sys.path.append('O:/Codes')
from credentials import proxy_request


proxies = proxy_request()

path = 'O:/Cruz/Inteligencia/Acompanhamento Global de Preços/Hog_Prices'
path_cattle = 'O:/Cruz'

'''
# Phase 1: Data collection
# - Collect historical hog prices from various countries from the Pig333 website.
# - Extract weekly hog prices from the European Union.
# - Collect weekly hog prices from the United Kingdom.
# - Extract hog price data for the US and Brazil from an Excel file.
# - Collect currency data from an Excel file.
'''
def get_swine_prices(country_code):
    url = 'https://www.pig333.com/markets_and_prices/?accio=cotitzacions'
    headers = {
        'Cookie': 'PHPSESSID=kbn40dm10len8pgeidie8td6gt; id_pais=8; seccio=markets_and_prices; sc=w_mob%20w_mobw%20w_tab%20w_des%20w_wid%20wmax_wid; __utmc=104705537; __utmz=104705537.1682510546.1.1.utmcsr=google|utmccn=(organic)|utmcmd=organic|utmctr=(not%20provided); uu=https%3A%2F%2Fwww.pig333.com%2Fmarkets_and_prices%2Fbelgium_158%2F; __utma=104705537.916163893.1682510546.1682510546.1682514515.2; __utmt=1; __utmt_b=1; __utmb=104705537.4.10.1682514515'
    }
    data = {
        'accio': 'cotitzacions',
        'moneda': 'USD',
        'unitats': 'cwt',
        'mercats': f'{country_code}'
    }

    response = requests.post(url, headers=headers, data=data, proxies=proxies, verify=False)

    df = pandas.DataFrame(response.json())
    i = 0
    for i in range(0, 9):
        df = df.drop(i)
        i += 1
    df = df.drop(columns=['taula'])
    df = df.drop(columns=['any_min'])
    df = df.drop(columns=['any_max'])
    df = df.drop(columns=['ids_mercats_ordre'])

    # df = df.head(1214)
    df['resultat']
    df['resultat'] = df['resultat'].str.replace('economia.data.addRow', '')
    df['resultat'] = df['resultat'].str.replace(',null,null', '')
    df['resultat'] = df['resultat'].str.replace(';', '')
    df['resultat'] = df['resultat'].str.replace('new Date', '')

    df.to_excel(f'{path}/SwinePrices_{country_code}.xlsx', index=False)

    wb = openpyxl.load_workbook(f"{path}/SwinePrices_{country_code}.xlsx")
    ws = wb.active

    i = 0
    for r in range(1, ws.max_row+1):
        for c in range(1, 2):
            s = ws.cell(r, c).value
            if s!=None and "," in s:
                ws.cell(r, c).value = s.replace(',', '/')
                i += 1
    i = 0
    for r in range(1, ws.max_row+1):
        for c in range(1, 2):
            s = ws.cell(r, c).value
            if s!=None and "([" in s:
                ws.cell(r, c).value = s.replace('([', '')
                i += 1
    i = 0
    for r in range(1, ws.max_row+1):
        for c in range(1, 2):
            s = ws.cell(r, c).value
            if s!=None and "/true])" in s:
                ws.cell(r, c).value = s.replace('/true])', '')
                i += 1
    i = 0
    for r in range(1, ws.max_row+1):
        for c in range(1, 2):
            s = ws.cell(r, c).value
            if s!=None and "])" in s:
                ws.cell(r, c).value = s.replace('])', '')
                i += 1
    i = 0
    for r in range(1, ws.max_row+1):
        for c in range(1, 2):
            s = ws.cell(r, c).value
            if s!=None and ")/" in s:
                ws.cell(r, c).value = s.replace(')/', ';')
                i += 1
    i = 0
    for r in range(1, ws.max_row+1):
        for c in range(1, 2):
            s = ws.cell(r, c).value
            if s!=None and "(" in s:
                ws.cell(r, c).value = s.replace('(', '')
                i += 1

    wb.save(f"{path}/SwinePrices_{country_code}.xlsx")

    df = pandas.read_excel(f"{path}/SwinePrices_{country_code}.xlsx")
    df = pandas.DataFrame(df)


    df[['Date', 'Prices']] = df['resultat'].str.split(';', expand=True)

    if country_code != 93:
        df = df.drop(0)
        df = df.drop(1)
    
    if country_code == 90:
        df = df.drop(columns=['resultat'])
        df = df.head(1320)
    else:
        df = df.drop(columns=['resultat'])
        df = df.head(510)
    
    df.to_excel(f"{path}/SwinePrices_{country_code}.xlsx", index=False)

    wb = openpyxl.load_workbook(f"{path}/SwinePrices_{country_code}.xlsx")
    ws = wb.active
    i = 0
    for r in range(1, ws.max_row+1):
        for c in range(2, 3):
            s = ws.cell(r, c).value
            if s!=None and "/" in s:
                ws.cell(r, c).value = s.replace('/', '')
                i += 1

    i = 0
    for i in range(1, ws.max_row+1):
        try:
            celula_b = ws[f'B{i}'].value
            partes = celula_b.split('.', 2)
            ws[f'B{i}'] = partes[0]+"."+partes[1]
        except:
            continue
        i += 1

    for r in range(1, ws.max_row+1):
        for c in range(2, 3):
            s = ws.cell(r, c).value
            if s!=None and "." in s:
                ws.cell(r, c).value = s.replace('.', ',')
                i += 1

    for r in range(1, ws.max_row+1):
        for c in range(2, 3):
            s = ws.cell(r, c).value
            if s!=None and "," in s:
                ws.cell(r, c).value = s[:5]
                i += 1

    wb.save(f"{path}/SwinePrices_{country_code}.xlsx")

countries = {
    'USA': 93,
    'FRA': 14,
    'BRA': 121,
    'CAN': 94,
    'MXN': 164,
    'KRW': 90,
    'COP': 148,
    'VND': 166,
    'CNY': 106,
    'RUS': 111
}

#Pull prices for Pig333 for Americas
can = get_swine_prices(94)
mxn = get_swine_prices(164)
cop = get_swine_prices(86)
col = get_swine_prices(148)
arg = get_swine_prices(149)

#Pull prices for Pig333 for Asia
cny = get_swine_prices(106)
skr = get_swine_prices(90)
jap = get_swine_prices(173)
vie = get_swine_prices(166)
twn = get_swine_prices(174)

#Pull prices for others
rus = get_swine_prices(111)

#Pull prices for European Union
url = 'https://agriculture.ec.europa.eu/document/download/8ecf65ae-208f-4630-8958-05bf13f09575_en?filename=pig-weekly-prices-eu_en_1.xlsx'

data_atual = datetime.datetime.now()
last_friday = data_atual - datetime.timedelta(days=(data_atual.weekday()+2)%7)
numero_semana = last_friday.isocalendar()[1]

response = requests.get(url, proxies=proxies, verify=False)
# response = requests.get(url, proxies=proxies)


df_main_europe = pandas.read_excel(response.content, sheet_name= 'Class E')
df_main_europe = pandas.DataFrame(df_main_europe)
i = 0
for i in range(0, 7):
    df_main_europe = df_main_europe.drop(i)
    i += 1
df_main_europe = df_main_europe.drop(8)
df_main_europe = df_main_europe.drop(columns = 'Unnamed: 28')
df_main_europe = df_main_europe.drop(columns = 'Unnamed: 8')
df_main_europe.columns = df_main_europe.iloc[0]
df_main_europe = df_main_europe[1:]
df_main_europe = df_main_europe[['E', 'BE', 'DK', 'DE', 'ES', 'FR', 'NL', 'PL', 'EU\n(weighted avg.)']]
df_main_europe = df_main_europe.drop(df_main_europe.tail(60-numero_semana+1).index) #Vai reduzindo assim se comer cotação na semana que vem
df_main_europe = df_main_europe.dropna()
df_main_europe = df_main_europe[::-1]
df_main_europe.columns = ['DATES', 'BELGIUM', 'DENMARK', 'GERMANY', 'SPAIN', 'FRANCE', 'NETHERLANDS', 'POLAND', 'EU WEIGHTED AVG']
df_main_europe['BELGIUM'] = pandas.to_numeric(df_main_europe['BELGIUM'], errors='coerce')
df_main_europe['DENMARK'] = pandas.to_numeric(df_main_europe['DENMARK'], errors='coerce')
df_main_europe['GERMANY'] = pandas.to_numeric(df_main_europe['GERMANY'], errors='coerce')
df_main_europe['SPAIN'] = pandas.to_numeric(df_main_europe['SPAIN'], errors='coerce')
df_main_europe['FRANCE'] = pandas.to_numeric(df_main_europe['FRANCE'], errors='coerce')
df_main_europe['NETHERLANDS'] = pandas.to_numeric(df_main_europe['NETHERLANDS'], errors='coerce')
df_main_europe['POLAND'] = pandas.to_numeric(df_main_europe['POLAND'], errors='coerce')
df_main_europe['EU WEIGHTED AVG'] = pandas.to_numeric(df_main_europe['EU WEIGHTED AVG'], errors='coerce')
df_main_europe['BELGIUM'] = df_main_europe['BELGIUM'].fillna(df_main_europe.iloc[1,1])
df_main_europe['DENMARK'] = df_main_europe['DENMARK'].fillna(df_main_europe.iloc[1,2])
df_main_europe['GERMANY'] = df_main_europe['GERMANY'].fillna(df_main_europe.iloc[1,3])
df_main_europe['SPAIN'] = df_main_europe['SPAIN'].fillna(df_main_europe.iloc[1,4])
df_main_europe['FRANCE'] = df_main_europe['FRANCE'].fillna(df_main_europe.iloc[1,5])
df_main_europe['NETHERLANDS'] = df_main_europe['NETHERLANDS'].fillna(df_main_europe.iloc[1,6])
df_main_europe['POLAND'] = df_main_europe['POLAND'].fillna(df_main_europe.iloc[1,7])
df_main_europe['EU WEIGHTED AVG'] = df_main_europe['EU WEIGHTED AVG'].fillna(df_main_europe.iloc[1,8])
df_main_europe.to_excel(f'{path}/EU_Hog_Prices_Weekly.xlsx', index = False)

#Pull prices for Great Britain
url = "https://projectblue.blob.core.windows.net/media/Default/Market%20Intelligence/pork/Prices/Deadweight%20pig%20prices/All%20historic%20pig%20data.xlsx"

data_atual = datetime.datetime.now()
last_friday = data_atual - datetime.timedelta(days=(data_atual.weekday()+2)%7)
numero_semana = last_friday.isocalendar()[1]

response = requests.get(url, proxies=proxies, verify=False)

df_main_great_britain = pandas.read_excel(response.content, sheet_name= 'Weekly')
df_main_great_britain = pandas.DataFrame(df_main_great_britain)
for i in range(0, 11):
    df_main_great_britain = df_main_great_britain.drop(i)
    i += 1
df_main_great_britain = df_main_great_britain.rename(columns={'Unnamed: 1': 'Date'})
df_main_great_britain = df_main_great_britain.rename(columns={'Unnamed: 3': 'Price (p/kg)'})
df_main_great_britain = df_main_great_britain[['Date', 'Price (p/kg)']]
df_main_great_britain = df_main_great_britain.dropna()
df_main_great_britain = df_main_great_britain.sort_values(by='Date',ascending=False)
df_main_great_britain['Date'] = pandas.to_datetime(df_main_great_britain['Date'], errors='coerce').dt.strftime('%d/%m/%Y')
df_main_great_britain['Price (p/kg)'] = pandas.to_numeric(df_main_great_britain['Price (p/kg)'], errors='coerce')
df_main_great_britain.to_excel(f'{path}/UK_Hog_Prices_Weekly.xlsx', index = False)

#Convert Base.xlsm and Base_currency.xlsm files
muda = openpyxl.load_workbook(f"{path_cattle}/Base.xlsm")
muda.save(f"{path}/Base.xlsx")
muda = openpyxl.load_workbook(f"{path_cattle}/Base_Currency.xlsm")
muda.save(f"{path}/Base_Currency.xlsx")