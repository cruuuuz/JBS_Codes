'''
# Objective: Collect and analyze historical hog prices globally.

# Techniques used:
# - Web Scraping: Extract data from websites like Pig333, European Union, and AHDB.
# - Data manipulation: Clean, process, and organize data with Pandas and Openpyxl.
# - Data analysis: Calculate averages and convert units of measure.
# - Data visualization: Create charts with Matplotlib and Mpld3.
# - Automation: Send emails with Outlook and update Snowflake database.


# Phase 3: Data analysis and visualization
# - Calculate weighted average prices for the European Union.
# - Create charts to visualize global price trends.
# - Generate separate charts for hog prices in the Americas and Asia.

# Phase 4: Automation
# - Send an email with a chart of global hog prices.
# - Update a Snowflake database with the collected and processed data.
'''

import pandas
import openpyxl
import sys
from pathlib import Path
from datetime import date
import snowflake.connector, snowflake.connector.pandas_tools as spd
sys.path.append('O:/Cruz/Codes/Inteligência')
from password import password
from snowflake_connector import snowflake_connector

proxies = {
    'http' : f'http://{password()}@MTZSVMFCPPRD02:8080',
    'https' : f'http://{password()}@MTZSVMFCPPRD02:8080'
}

snowflake_connector = snowflake_connector()
cnn = snowflake.connector.connect(
    user = snowflake_connector['user'],
    password = snowflake_connector['password'],
    account = snowflake_connector['account'],
    proxy_host = snowflake_connector['proxy_host'],
    proxy_port = snowflake_connector['proxy_port']
)
cursor = cnn.cursor()

path = 'O:/Cruz/Inteligencia/Acompanhamento Global de Preços/Hog_Prices'
path_cattle = 'O:/Cruz'



'''
# Phase 2: Data cleaning and processing
# - Clean and format the collected data, including dates, prices, and units of measure.
# - Convert currencies to USD/cwt.
# - Handle missing and inconsistent values.
'''

#Pull currencies 
wb_currency = openpyxl.load_workbook(f"{path}/Base_Currency.xlsx")
ws_currency = wb_currency['Hist']


#Parse Canada data
wb_canada = openpyxl.load_workbook(f"{path}/SwinePrices_94.xlsx")
ws_canada = wb_canada.active

df_can = pandas.DataFrame(ws_canada.values)
df_can.columns = [['Date', 'CAN']]
df_can.set_axis = df_can.iloc[0]
df_can = df_can.drop(0)

#Parse China data
wb_china = openpyxl.load_workbook(f"{path}/SwinePrices_94.xlsx")
ws_china = wb_china.active

df_cny = pandas.DataFrame(ws_china.values)
df_cny.columns = [['Date', 'CNY']]
df_cny.set_axis = df_cny.iloc[0]
df_cny = df_cny.drop(0)

#Parse Mexico data
wb_mexico = openpyxl.load_workbook(f"{path}/SwinePrices_164.xlsx")
ws_mexico = wb_mexico.active

df_mex = pandas.DataFrame(ws_mexico.values)
df_mex.columns = [['Date', 'MEX']]
df_mex.set_axis = df_mex.iloc[0]
df_mex = df_mex.drop(0)

#Bringing countries together
file_eur = f"{path}/EU_Hog_Prices_Weekly.xlsx"
wb_eur = openpyxl.load_workbook(file_eur)
ws_eur = wb_eur.active
df_eur = pandas.DataFrame(ws_eur.values)
df_eur = df_eur.drop(columns=[0])
df_eur = df_eur.head(1200)

file_uk = f"{path}/UK_Hog_Prices_Weekly.xlsx"
wb_uk = openpyxl.load_workbook(file_uk)
ws_uk = wb_uk.active
df_uk = pandas.DataFrame(ws_uk.values)
df_uk = df_uk.drop(columns=[0])
df_uk = df_uk.head(1200)

wb_bbg = openpyxl.load_workbook(f"{path}/Base.xlsx")
ws_bbg = wb_bbg['Base_Hog_PricesUSD']
df_usa_bra = pandas.DataFrame(ws_bbg.values)
df_usa_bra.columns = df_usa_bra.iloc[0]
df_usa_bra['Date'] = pandas.to_datetime(df_usa_bra['Date'], errors='coerce').dt.strftime('%d/%m/%Y')
df_usa_bra = df_usa_bra.head(1200)

df_global = pandas.concat([df_usa_bra, df_eur], axis=1)
df_global = pandas.concat([df_global, df_uk], axis=1)
df_global['Date'] = pandas.to_datetime(df_global['Date'], errors='coerce').dt.strftime('%d/%m/%Y') 
df_global.to_excel(f"{path}/Temp_Global_Swine_Prices.xlsx", index=False, header=False)

wb_temp_base = openpyxl.load_workbook(f"{path}/Temp_Global_Swine_Prices.xlsx")
ws_temp_base = wb_temp_base.active

for row in ws_temp_base.iter_rows():
    for cell in row:
        if isinstance(cell.value, str or int) and "(*)" in cell.value:
            cell.value = cell.value.replace("(*)", '')
for r in range(1, ws_temp_base.max_row+1):
    for c in range(1, ws_temp_base.max_column+1):
        s = ws_temp_base.cell(r,c).value
        if s == None:
            ws_temp_base.cell(r,c).value = 0
wb_temp_base.save(f"{path}/Temp_Global_Swine_Prices.xlsx")

#Convert measure units
wb_base = openpyxl.load_workbook(f"{path}/Temp_Global_Swine_Prices.xlsx")
ws_base = wb_base.active

wb_canada = openpyxl.load_workbook(f"{path}/SwinePrices_94.xlsx")
ws_canada = wb_canada.active

wb_mexico = openpyxl.load_workbook(f"{path}/SwinePrices_164.xlsx")
ws_mexico = wb_mexico.active

wb_china = openpyxl.load_workbook(f"{path}/SwinePrices_106.xlsx")
ws_china = wb_china.active

wb_russia = openpyxl.load_workbook(f"{path}/SwinePrices_111.xlsx")
ws_russia = wb_russia.active

linha = 1
for i in range(1, 1200):
    try:
        ws_base[f"C{linha+1}"] = ws_base[f"C{linha+1}"].value/ws_currency[f"C{linha+1}"].value/2.2046/0.775*100 #Convering BR to USD/cwt and dressed
        ws_base[f"D{linha+1}"] = ws_base[f"D{linha+1}"].value*ws_currency[f"D{linha+1}"].value/2.2046 #Converting Belgium to USD/cwt
        ws_base[f"E{linha+1}"] = ws_base[f"E{linha+1}"].value*ws_currency[f"D{linha+1}"].value/2.2046 #Converting Denmark to USD/cwt
        ws_base[f"F{linha+1}"] = ws_base[f"F{linha+1}"].value*ws_currency[f"D{linha+1}"].value/2.2046 #Converting Germnay to USD/cwt
        ws_base[f"G{linha+1}"] = ws_base[f"G{linha+1}"].value*ws_currency[f"D{linha+1}"].value/2.2046 #Converting Spain to USD/cwt
        ws_base[f"H{linha+1}"] = ws_base[f"H{linha+1}"].value*ws_currency[f"D{linha+1}"].value/2.2046 #Converting France to USD/cwt
        ws_base[f"I{linha+1}"] = ws_base[f"I{linha+1}"].value*ws_currency[f"D{linha+1}"].value/2.2046 #Converting Netherlands to USD/cwt
        ws_base[f"J{linha+1}"] = ws_base[f"J{linha+1}"].value*ws_currency[f"D{linha+1}"].value/2.2046 #Poland Germnay to USD/cwt
        ws_base[f"N1"] = "EU AVG"
        ws_base[f"N{linha+1}"] = ws_base[f"K{linha+1}"].value*ws_currency[f"D{linha+1}"].value/2.2046 #Converting EU Avg to USD/cwt
        ws_base[f"O1"] = "UK"
        ws_base[f"O{linha+1}"] = ws_base[f"L{linha+1}"].value*ws_currency[f"L{linha+1}"].value/2.2046 #Converting UK to USD/cwt
        ws_base[f"K1"] = "CAN"
        ws_base[f"L1"] = "MEX"
        ws_base[f"M1"] = "CHINA"
        ws_base[f"P1"] = "RUSSIA"
        ws_base[f"K{linha+1}"] = float(str(ws_canada[f"B{linha+1}"].value).replace(',','.'))/0.75 #Canada values
        ws_base[f"L{linha+1}"] = str(ws_mexico[f"B{linha+1}"].value).replace(',','.')
        ws_base[f"L{linha+1}"] = float(str(ws_base[f"L{linha+1}"].value).replace('None','100'))/0.791 #Converting Mexico (already in USD/cwt) to dressed
        ws_base[f"M{linha+1}"] = float(str(ws_china[f"B{linha+1}"].value).replace(',','.'))/0.75 #Converting China (already in USD/cwt) to dressed
        ws_base[f"P{linha+1}"] = float(str(ws_russia[f"B{linha+1}"].value).replace(',','.'))/0.76 #Converting Russia (already in USD/cwt) to dressed
    except Exception as e:
        print(f'error while data was being processed: {e}')
    linha += 1

df_swine_prices = pandas.DataFrame(ws_base.values)
df_swine_prices = df_swine_prices.head(1200)
df_swine_prices = df_swine_prices.dropna()
df_swine_prices.iloc[0,0] = 'Date'
df_swine_prices.columns = df_swine_prices.loc[0]
df_swine_prices['Date'] = pandas.to_datetime(df_swine_prices['Date'], errors='coerce').dt.strftime('%d/%m/%Y')
df_swine_prices.iloc[0,0] = 'Date'
df_swine_prices.to_csv(f"{path}/Global_Swine_Prices.csv", index=False, header=False, sep=';')
df_swine_prices.to_excel(f"{path}/Base_Global_Swine_Prices.xlsx", index=False, header=False)









